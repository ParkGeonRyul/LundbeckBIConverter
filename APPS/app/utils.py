import pandas as pd
import shutil
import os

from openpyxl import load_workbook
from datetime import datetime
from app.classes import *


def back_up(file_path: str):
    current_datetime = datetime.now().strftime('%Y%m%d%H%M')
    base_file_name = os.path.basename(file_path)
    backup_file_name = f"{os.path.splitext(base_file_name)[0]}_{current_datetime}.xlsx"
    backup_file_path = os.path.join(backup_folder, backup_file_name)
    
    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)
    
    shutil.copy(file_path, backup_file_path)

def remove_sheets(file_path: str, bi_sheet_name: str):
    try:
        workbook = load_workbook(file_path)
        if bi_sheet_name in workbook.sheetnames:
            std = workbook[bi_sheet_name]
            workbook.remove(std)
            workbook.save(file_path)
        workbook.close()
    except FileNotFoundError:
        print("파일을 찾을 수 없습니다. 경로를 확인해 주세요.")
        raise 

def data_cycles(sheet_name: str, df: pd.DataFrame, melted_column: list):
    all_data = pd.DataFrame()

    if sheet_name == CapacityClass().sheet_name:
        testa = 6
        column_value = 5

    elif sheet_name == PromotionClass().sheet_name:
        testa = 7
        column_value = 6

    elif sheet_name == PcrClass().sheet_name:
        column_value = 6

    fixed_cols = df.columns[:column_value]
    cycles = []

    if sheet_name != PcrClass().sheet_name:

        for i in range(4):
            increase_count = testa + (i * 13)
            range_count = {'range': range(increase_count, increase_count + 12), 'qetable': f'QE{i + 1}'}
            cycles.append(range_count)

    else:
        cycles = PcrClass().cycles

    for cycle in cycles:
        cycle_range = cycle['range']
        qetable_value = cycle['qetable']        
        date_cols = df.columns[cycle_range]      
        melted = df.melt(id_vars=fixed_cols, value_vars=date_cols, var_name='DATES', value_name='VALUE')
        melted.columns = melted_column

        if sheet_name == PcrClass().sheet_name:
            melted['VALUE'] = melted['VALUE'].apply(lambda x: 0 if x == '-' else x)

        melted['QETABLE'] = qetable_value
        all_data = pd.concat([all_data, melted], ignore_index=True)
    
    return all_data