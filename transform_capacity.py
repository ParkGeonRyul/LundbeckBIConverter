import pandas as pd
from openpyxl import load_workbook

# 함수 정의: 이 함수는 Excel 파일에서 Capacity 시트를 변환하여 Capacity_BI 시트에 저장합니다.
def transform_capacity_to_pivot(file_path, capacity_sheet_name='Capacity', capacity_bi_sheet_name='Capacity_BI'):
    # Excel 파일에서 Capacity 시트를 로드합니다.
    capacity_df = pd.read_excel(file_path, sheet_name=capacity_sheet_name)
    
    # 기존 Capacity_BI 시트가 존재하면 삭제합니다.
    try:
        workbook = load_workbook(file_path)
        if capacity_bi_sheet_name in workbook.sheetnames:
            std = workbook[capacity_bi_sheet_name]
            workbook.remove(std)
            workbook.save(file_path)
        workbook.close()
    except FileNotFoundError:
        print("파일을 찾을 수 없습니다. 경로를 확인해 주세요.")
        return
    
    # D열에서 "ZGR"로 시작하는 행 제외
    capacity_df = capacity_df[~capacity_df.iloc[:, 3].astype(str).str.startswith('ZGR', na=False)]

    # C열에서 LEN(TRIM(C열)) = 10 값만 필터링
    # 실제로는 8이 맞으나 파이썬에서 가지고 올때 .0 포함한다. 그래서 8이 아닌 10으로 변환함.
    capacity_df = capacity_df[capacity_df.iloc[:, 2].astype(str).str.strip().str.len() == 10]

    # 필요한 고정 열을 추출합니다 (A~E열).
    fixed_cols = capacity_df.columns[:5]
    
    # 변환된 데이터를 저장할 빈 데이터프레임을 만듭니다.
    all_data = pd.DataFrame()
    
    # 사이클별로 데이터를 처리합니다.
    cycles = [
        {'range': range(6, 18), 'qetable': 'QE1'},   # H~S 열의 인덱스 범위
        {'range': range(19, 31), 'qetable': 'QE2'},  # U~AC 열의 인덱스 범위
        {'range': range(32, 44), 'qetable': 'QE3'},  # AD~AO 열의 인덱스 범위
        {'range': range(45, 57), 'qetable': 'QE4'}   # AP~BA 열의 인덱스 범위
    ]
    
    for cycle in cycles:
        cycle_range = cycle['range']
        qetable_value = cycle['qetable']
        
        # 현재 사이클의 열 이름을 추출합니다.
        date_cols = capacity_df.columns[cycle_range]
        
        # 데이터를 피벗 형식으로 변환합니다.
        capacity_melted = capacity_df.melt(id_vars=fixed_cols, value_vars=date_cols, var_name='DATES', value_name='VALUE')
        
        # Capacity_BI 형식에 맞게 열 이름을 변경합니다.
        capacity_melted.columns = ['Capacity', 'Cost_Center', 'Code', 'Account group', 'Versions', 'DATES', 'VALUE']
        
        # 'QETABLE' 컬럼 추가
        capacity_melted['QETABLE'] = qetable_value  # 하드코딩된 QETABLE 값 사용
        
        # 변환된 데이터를 모두 합칩니다.
        all_data = pd.concat([all_data, capacity_melted], ignore_index=True)
    
    # 변환된 데이터를 Capacity_BI 시트에 저장합니다.
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        all_data.to_excel(writer, sheet_name=capacity_bi_sheet_name, index=False)
    
    # I1 셀에 'USED' 텍스트 추가
    workbook = load_workbook(file_path)
    sheet = workbook[capacity_bi_sheet_name]
    sheet['I1'] = 'USED'
    workbook.save(file_path)
    workbook.close()
    
    print(f"데이터가 변환되어 {capacity_bi_sheet_name} 시트에 저장되었습니다.")

# 파일 경로와 시트 이름을 정의합니다.
file_path = 'C:/WORKING/2024 QE Data (SG&A)_r1_20240716.xlsx'
capacity_sheet_name = 'Capacity'
capacity_bi_sheet_name = 'Capacity_BI'

# 변환 함수 실행
transform_capacity_to_pivot(file_path, capacity_sheet_name, capacity_bi_sheet_name)
