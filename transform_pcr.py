import pandas as pd
from openpyxl import load_workbook
import shutil
import os
from datetime import datetime
 
# 함수 정의
def transform_PCR_Power_BI_to_pivot(file_path, pcr_power_sheet_name='PCR_Power', pcr_power_bi_sheet_name='PCR_POWERBI'):
    # 현재 날짜와 시간을 가져와서 백업 파일 이름 생성
    current_datetime = datetime.now().strftime('%Y%m%d%H%M')
    backup_folder = 'c:/WORKING/BACKUP'
    backup_file_path = os.path.join(backup_folder, f'2024 PCR DATA_{current_datetime}.xlsx')
    # 백업 폴더가 없으면 생성
    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)
    # 원본 파일을 백업 폴더로 복사
    shutil.copy(file_path, backup_file_path)
 
    # 데이터 변환 작업 시작
    pcr_power_df = pd.read_excel(file_path, sheet_name=pcr_power_sheet_name)
    try:
        workbook = load_workbook(file_path)
        if pcr_power_bi_sheet_name in workbook.sheetnames:
            std = workbook[pcr_power_bi_sheet_name]
            workbook.remove(std)
            workbook.save(file_path)
        workbook.close()
    except FileNotFoundError:
        print("파일을 찾을 수 없습니다. 경로를 확인해 주세요.")
        return
    fixed_cols = pcr_power_df.columns[:5]
    all_data = pd.DataFrame()
    cycles = [
        {'range': range(11, 23), 'qetable': 'FY ACT 2023 @BUD rate'},
        {'range': range(24, 36), 'qetable': 'FY BUD 2024 @BUD rate'},
        {'range': range(37, 49), 'qetable': '24 QE1'},
        {'range': range(50, 62), 'qetable': '24 QE2'},
        {'range': range(63, 75), 'qetable': '24 QE3'},
        {'range': range(76, 88), 'qetable': '24 QE4'},
        {'range': range(89, 101), 'qetable': 'YTD ACT 2024 @BUD rate'},
        {'range': range(102,114 ), 'qetable': 'FY BUD 2025 @BUD rate'}
    ]
    for cycle in cycles:
        cycle_range = cycle['range']
        qetable_value = cycle['qetable']
        date_cols = pcr_power_df.columns[cycle_range]
        pcr_power_melted = pcr_power_df.melt(id_vars=fixed_cols, value_vars=date_cols, var_name='DATES', value_name='VALUE')
 
        # VALUE 값이 '-'인 경우 0으로 변경
        pcr_power_melted['VALUE'] = pcr_power_melted['VALUE'].apply(lambda x: 0 if x == '-' else x)
 
        pcr_power_melted['QETABLE'] = qetable_value
        pcr_power_melted.columns = ['COA', 'Account Group','Function1#','Product Grp','Category4', 'DATES', 'VALUE', 'QETABLE']
        all_data = pd.concat([all_data, pcr_power_melted], ignore_index=True)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        all_data.to_excel(writer, sheet_name=pcr_power_bi_sheet_name, index=False)
    # Open the workbook again to add "Category5" and "USED" in I1 and J1
    workbook = load_workbook(file_path)
    sheet = workbook[pcr_power_bi_sheet_name]
    sheet['I1'] = 'Category5'
    sheet['J1'] = 'USED'
    workbook.save(file_path)
    print(f"데이터가 변환되어 {pcr_power_bi_sheet_name} 시트에 저장되었습니다.")
 
# 파일 경로와 시트 이름을 정의합니다.
file_path = 'c:/WORKING/2024 PCR DATA.xlsx'
pcr_power_sheet_name = 'PCR_Power'
pcr_power_bi_sheet_name = 'PCR_POWERBI'
 
# 변환 함수 실행
transform_PCR_Power_BI_to_pivot(file_path, pcr_power_sheet_name, pcr_power_bi_sheet_name)