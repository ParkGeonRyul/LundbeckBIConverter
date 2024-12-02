import pandas as pd
from openpyxl import load_workbook
import shutil
import os
from datetime import datetime

# 함수 정의
def transform_promotion_to_pivot(file_path, promotion_sheet_name='Promotion', promotion_bi_sheet_name='Promotion_BI'):
    # 현재 날짜와 시간을 가져와서 백업 파일 이름 생성
    current_datetime = datetime.now().strftime('%Y%m%d%H%M')
    backup_folder = 'c:/WORKING/BACKUP'
    base_file_name = os.path.basename(file_path)
    backup_file_name = f"{os.path.splitext(base_file_name)[0]}_{current_datetime}.xlsx"
    backup_file_path = os.path.join(backup_folder, backup_file_name)
    
    # 백업 폴더가 없으면 생성
    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)
    
    # 원본 파일을 백업 폴더로 복사
    shutil.copy(file_path, backup_file_path)
    # print(f"백업 파일이 생성되었습니다: {backup_file_path}")

    # 데이터 변환 작업 시작
    promotion_df = pd.read_excel(file_path, sheet_name=promotion_sheet_name)
    
    try:
        workbook = load_workbook(file_path)
        if promotion_bi_sheet_name in workbook.sheetnames:
            std = workbook[promotion_bi_sheet_name]
            workbook.remove(std)
            workbook.save(file_path)
        workbook.close()
    except FileNotFoundError:
        print("파일을 찾을 수 없습니다. 경로를 확인해 주세요.")
        return
    
    fixed_cols = promotion_df.columns[:6]
    all_data = pd.DataFrame()
    
    cycles = [
        {'range': range(7, 19), 'qetable': 'QE1'},
        {'range': range(20, 32), 'qetable': 'QE2'},
        {'range': range(33, 45), 'qetable': 'QE3'},
        {'range': range(46, 58), 'qetable': 'QE4'}
    ]
    
    for cycle in cycles:
        cycle_range = cycle['range']
        qetable_value = cycle['qetable']
        date_cols = promotion_df.columns[cycle_range]
        
        promotion_melted = promotion_df.melt(id_vars=fixed_cols, value_vars=date_cols, var_name='DATES', value_name='VALUE')
        promotion_melted.columns = ['PROMOTION', 'CODE', 'CODE_NAME', 'PRODUCT_GROUP', 'PRODUCT_CODE', 'VERSION', 'DATES', 'VALUE']
        promotion_melted['QETABLE'] = qetable_value
        
        all_data = pd.concat([all_data, promotion_melted], ignore_index=True)
    
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        all_data.to_excel(writer, sheet_name=promotion_bi_sheet_name, index=False)
    
    # Open the workbook again to add "USED" in J1
    workbook = load_workbook(file_path)
    sheet = workbook[promotion_bi_sheet_name]
    sheet['J1'] = 'USED'
    workbook.save(file_path)
    
    print(f"데이터가 변환되어 {promotion_bi_sheet_name} 시트에 저장되었습니다.")

# 파일 경로와 시트 이름을 정의합니다.
file_path = 'c:/WORKING/2024 QE Data (SG&A).xlsx'
promotion_sheet_name = 'Promotion'
promotion_bi_sheet_name = 'Promotion_BI'

# 변환 함수 실행
transform_promotion_to_pivot(file_path, promotion_sheet_name, promotion_bi_sheet_name)
